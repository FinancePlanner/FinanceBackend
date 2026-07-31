"""Generate messy real-world-ish .xlsx fixtures for the CoreXLSX spike."""
import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fixtures")
os.makedirs(OUT, exist_ok=True)


def messy_workbook(path):
    """Header not in row 1, data starting at column G, a total row, formulas,
    a second non-expense sheet, and a third sheet with a different layout."""
    wb = Workbook()

    # --- Sheet 1: the realistic mess -------------------------------------
    ws = wb.active
    ws.title = "Gastos 2026"
    ws["B2"] = "Orçamento Pessoal"
    ws["B2"].font = Font(bold=True, size=16)
    ws["B3"] = "atualizado em"
    ws["C3"] = dt.date(2026, 7, 1)

    # header row is row 5, data region starts at column G
    headers = ["Data", "Descrição", "Categoria", "Valor", "Notas"]
    for i, h in enumerate(headers):
        c = ws.cell(row=5, column=7 + i, value=h)
        c.font = Font(bold=True)

    rows = [
        (dt.date(2026, 1, 3), "Continente", "Supermercado", 84.20, ""),
        (dt.date(2026, 1, 5), "Netflix", "Lazer", 13.99, "mensal"),
        (dt.date(2026, 1, 11), "Renda", "Casa", 750.00, ""),
        (dt.date(2026, 2, 2), "Galp", "Transportes", 61.40, ""),
        (dt.date(2026, 2, 14), "Jantar fora", "Lazer", 45.00, "aniversário"),
        (dt.date(2026, 3, 1), "Renda", "Casa", 750.00, ""),
        (dt.date(2026, 3, 9), "Farmácia", "Saúde", 22.15, ""),
        (dt.date(2026, 3, 22), "PingoDoce", "Supermercado", 103.77, ""),
    ]
    r = 6
    for d, desc, cat, val, note in rows:
        ws.cell(row=r, column=7, value=d).number_format = "DD/MM/YYYY"
        ws.cell(row=r, column=8, value=desc)
        ws.cell(row=r, column=9, value=cat)
        ws.cell(row=r, column=10, value=val).number_format = '#,##0.00 "€"'
        ws.cell(row=r, column=11, value=note)
        r += 1

    # the trap: a total row written as a formula, plus a stray label
    ws.cell(row=r, column=9, value="TOTAL").font = Font(bold=True)
    total = ws.cell(row=r, column=10, value=f"=SUM(J6:J{r - 1})")
    total.number_format = '#,##0.00 "€"'
    total.font = Font(bold=True)

    # a derived column that is a formula over another column
    ws.cell(row=5, column=12, value="Valor c/ IVA").font = Font(bold=True)
    for i in range(6, r):
        ws.cell(row=i, column=12, value=f"=J{i}*1.23")

    # --- Sheet 2: not expense data at all --------------------------------
    ws2 = wb.create_sheet("Resumo")
    ws2["A1"] = "Pilar"
    ws2["B1"] = "Alvo %"
    for i, (name, pct) in enumerate(
        [("Essenciais", 50), ("Futuro", 30), ("Lazer", 20)], start=2
    ):
        ws2.cell(row=i, column=1, value=name)
        ws2.cell(row=i, column=2, value=pct)

    # --- Sheet 3: clean-ish, header row 1, text dates --------------------
    ws3 = wb.create_sheet("2025 (antigo)")
    for i, h in enumerate(["date", "merchant", "amount", "pillar"], start=1):
        ws3.cell(row=1, column=i, value=h)
    old = [
        ("12/11/2025", "Uber", 9.30, "fun"),
        ("13/11/2025", "Lidl", 41.05, "fundamentals"),
        ("28/11/2025", "Ginásio", 34.90, "fun"),
    ]
    for j, (d, m, a, p) in enumerate(old, start=2):
        ws3.cell(row=j, column=1, value=d)  # text date, DD/MM/YYYY
        ws3.cell(row=j, column=2, value=m)
        ws3.cell(row=j, column=3, value=a)
        ws3.cell(row=j, column=4, value=p)

    wb.save(path)


def formulas_without_cache(path):
    """openpyxl never writes cached formula values -- this is exactly the
    'formula cell with no cached value' case the plan calls out."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["date", "title", "amount"])
    ws.append([dt.date(2026, 5, 1), "Rent", 700])
    ws.append([dt.date(2026, 5, 2), "Coffee", "=3.5*2"])
    wb.save(path)


def wide(path, rows=5000):
    wb = Workbook()
    ws = wb.active
    ws.append(["date", "title", "amount", "category"])
    base = dt.date(2024, 1, 1)
    for i in range(rows):
        ws.append([base + dt.timedelta(days=i % 900), f"item {i}", 10 + (i % 97) / 3, f"cat{i % 12}"])
    wb.save(path)


messy_workbook(os.path.join(OUT, "messy.xlsx"))
formulas_without_cache(os.path.join(OUT, "no-cached-formula.xlsx"))
wide(os.path.join(OUT, "wide-5000.xlsx"))
print("wrote:", sorted(os.listdir(OUT)))
